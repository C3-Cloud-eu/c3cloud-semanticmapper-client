import yaml
import json
import sys
import os.path
from openpyxl import load_workbook
import re
import requests
import unicodedata
from collections import Counter
import time
import pandas as pd

# assumptions about the xlsx structure:
# [Concepts][        site name 1       ][        site name 2       ]...
# [        ][codesys][code][designation][codesys][code][designation]...
# ....rows....
#
# the first column of each "site name X" can be anything and have to be specified
# in the YAML config file
#
# refer to the documentation for more details


# baseurl = sys.argv[0]
#baseurl = 'http://localhost:5000/c3-cloud/'  # tests
# baseurl = 'http://cispro.chu-rouen.fr/c3-cloud/'
# baseurl = "http://18.205.143.209/c3-cloud/"

# colored terminal output
class bcolors:
    OKBLUE = '\033[94m'
    OKGREEN = '\033[92m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'


# helper function: print with a specified color
def printcolor(s, color, end='\n'):
    print(color, end='')
    print(s, end='')
    print(bcolors.ENDC, end=end)

def printinfo(s):
    printcolor(s, color=bcolors.OKBLUE)
def printok(s):
    printcolor(s, color=bcolors.OKGREEN)
def printwarn(s):
    printcolor(s, color=bcolors.WARNING)
def printerr(s):
    printcolor(s, color=bcolors.FAIL)

# constants
SITE = 'site'
CODE_SYSTEM = 'code_system'
CODE_SYSTEM_URI = 'code_system_uri'
CODE_SYSTEM_VERSION = 'code_system_version'
CODE = 'code'
DESIGNATION = 'designation'
CONCEPT = 'concept'
EQUIVALENCE = 'equivalence'
EQUIVALENCE_DEFINITION = 'equivalence_definition'


# ### requests ### #

# normalize and trim a string
def clean(s):
    if s is None:
        return None
    s = str(s)
    if 'TBC by' in s:
        s = ''
    return(unicodedata.normalize('NFKC', str(s)).strip())


# send a request to the REST API
# url: the part of the URL coming *after* the baseurl
# method: "get" or "post"
# data: json encoded data to send as the POST request body
# get: GET parameters
def sendrequest(url, method="get", get=None, data=None):
    full_url = "{}{}/".format(baseurl, url)
    args = {k: v for k, v in {'params': get, 'json': data}.items() if v}
    # execute either requests.get or requests.post based on the <method> arg
    r = getattr(requests, method)(url=full_url, **args)
    if interactive:
        print(method)
        print(full_url)
        print(args)
        print(r)
        print(r.status_code)
    if r.status_code != 200:
        report['error'] += 1
        print("error", r.status_code, r.text)
    return(r.text)


# send a mapping to upload
# o: python dict containing the mapping data
def upload_mapping(o):
    printcolor('[uploading <{}>@<{}>]'.format(o[CONCEPT], o[SITE]),
               color=bcolors.OKBLUE, end='')
    print(o)
    printwarn('----')
    if len(o['codes']) > 0:
        codesystem = o['codes'][0][CODE_SYSTEM]
        try:
            # lookup the code system uri from the server
            uri = [cs[CODE_SYSTEM_URI] for cs in codesystems if cs[CODE_SYSTEM] == codesystem][0]
        except IndexError:
            printcolor('no such URI for {}'.format(codesystem), color=bcolors.FAIL)
            report['error'] += 1
            return None

        # add the uri to the POST data to send
        def f(c):
            c[CODE_SYSTEM_URI] = uri
            return(c)
        o['codes'] = [f(c) for c in o['codes']]
    print(o)
    quit()
    if not dryrun:
        ans = sendrequest("mappings", method="post", data=o)
        print(ans)


# ### model ### #

# given the dictionary of {concept: [mapping]},
# build the object to send as json data,
# check wether it already exist and decide to upload it or not
def process_items(d):
    # # sorted lists for objects comparison
    def f(l):
        return sorted( [set([e[k] for k in [CODE, DESIGNATION, CODE_SYSTEM]]) for e in l] )
    #return(sorted([[v for k, v in x.items() if k in [CODE, DESIGNATION, CODE_SYSTEM]]
        #              for x in l]))

    # "pretty print"
    def disp(title, l):
        print('\n' +
              '\n'.join([title+':'] + [str([e, i[e]]) for i in l for e in sorted(i.keys())]) +
              '\n')
    
    for concept, sites in d.items():
        for site, items in sites.items():
            # print("processing ", concept, " -- ", site, " -- ", items)

            # time.sleep(0.5)
            # items = [{clean(k): clean(v) for k, v in i.items()} for i in items]
            o = {
                'concept': clean(concept),
                'site': clean(site),
                'codes': [{clean(i): clean(item[i]) for i in [CODE_SYSTEM, CODE, DESIGNATION]}
                          for item in items]}
            # disp("test", [o])
            # get the mappings that already exist on the server
            codelist = [i for i in mappings if i[CONCEPT] == o['concept']
                        and i[SITE] == o['site']]
            if len(codelist) > 0:
                print("already in db→ <{}>@<{}>:".format(concept, site), end='')
                # print('<<<<<<<<',f(o['codes']),'||||||||||||||||',f(codelist),'>>>>>>>')

                if f(o['codes']) == f(codelist):
                # if compare(o['codes'], codelist):
                    printcolor("[identical]", color=bcolors.OKGREEN, end='')
                    report['identical'] += 1
                else:
                    printcolor("[different]", color=bcolors.WARNING, end='')
                    disp('local', items)
                    disp('server', codelist)
                    print('to upload: {}', o)
                    report['different'] += 1

                    if(FORCE):
                        upload_mapping(o)
                    else:
                        printcolor('[use --force to overwrite]', color=bcolors.FAIL, end='')
            else:
                upload_mapping(o)
                print([e[CODE] for e in items], end='')
                report['new'] += 1
            print('')
            if interactive:
                print(">>>>", o)
                input()



# ### xlsx ### #


# coordinate manipulation
class Coord:
    def __init__(self, col, row):
        self.col = col
        self.row = row

    def __str__(self):
        return self.col + str(self.row)

    def __repr__(self):
        return self.__str__()

    def horizontal(self, dir=1):
        return Coord(chr(ord(self.col) + dir), self.row)

    def vertical(self, dir=1):
        return Coord(self.col, self.row + dir)

    def down(self, n=1):
        return self.vertical(n)
    
    def right(self, n=1):
        return self.horizontal(n)

    def left(self, n=1):
        return self.horizontal(-n)

    def top(coord):  # get the associated site
        return Coord(coord.col, 1)

    def concept(coord):  # get the associated concept
        return Coord("A", coord.row)


# retrieve a reference to another cell
# ie if the value of the cell is "=F42", retrieve the value at F42 in the current sheet
# and if it is "=$'some other sheet'.=F42", retrieve the F42 value in the sheet "some other sheet" in the workbook
def resolveCell(wb, sheet, coord):
    if isinstance(coord, Coord):
        coord = str(coord)
    # print("<<<<<<<{}>>>>>>>>".format(coord), type(coord))
    
    v = sheet[coord].value
    if isinstance(v, str) and v[0] == '=':
        if '.' in v:
            sht, coord = v.split('.')
            sht = wb[re.sub(r"^\$'|'$", "", sht)]
        else:
            sht = sheet
            coord = v
        coord = coord[1:]
        return(clean(sht[coord].value))
    else:
        # return('' if v is None else v)
        return(clean(v))


def build_item(wb, sheet, startcoord):
    # headerLeft = startcoord.top().down()
    # headers = [ resolveCell(wb, sheet, startcoord.top().down().right(i)) for i in range(0,2) ]

    def cell(c):
        return resolveCell(wb, sheet, str(c))
    
    def findcol(col):
        
        # print("--------findcol "+col)
        for i in range(0,3):
            c = startcoord.right(i).top().down()
            # print(c)
            cv = cell(c)
            if clean(cv) == clean(col):
                # print('col found: ',c,cv)
                # printinfo(cell(startcoord.right(i)))
                return cell(startcoord.right(i))
            # else:
            #     print("no")
    # # print("building item for {}".format(startcoord),
    # #       sheet[str(startcoord)].value)
    # coords = {
    #     SITE: startcoord.top(),
    #     CODE_SYSTEM: startcoord,
    #     CODE:        startcoord.right(),
    #     DESIGNATION: startcoord.right(2),
    #     CONCEPT:     startcoord.concept()}
    item = {
        SITE:        cell(startcoord.top()),
        CODE_SYSTEM: findcol("Coding System"),
        CODE:        findcol("Code"),
        DESIGNATION: findcol("Designation"),
        CONCEPT:     cell(startcoord.concept())}

    # allow empty designation
    if item[DESIGNATION] is None:
        printerr('{} {} is empty'.format(item[SITE], item[CODE]))
        item[DESIGNATION] = ''
    # print(item)
    # print(cell(startcoord.top()), cell(startcoord.concept()))

    # return({k: str(resolveCell(wb, sheet, str(v)))
    #         for k, v in coords.items()})
    assert(item[DESIGNATION] is not None)
    if any([v is None for k, v in item.items() if
            k == CODE
            or k == CODE_SYSTEM
            or k == SITE
            or k == CONCEPT]) :
        return None
    
    if any([v is None for k, v in item.items() if k == CODE_SYSTEM or k == CODE or k == CONCEPT or k == SITE]) :
        printcolor("$$ {} --> {} ({} -- {}) ".format(startcoord,
                                              item,
                                              wb, sheet), color=bcolors.FAIL)
        # assert False
    return item

def item_is_valid(i):
    # printinfo('checking {}'.format(i))
    def is_illegal(s):
        s = str(s).lower()
        # print(i, s)
        illegal = (s is None
                   or s in ['', 'no coding', 'no code', 'no designation', 'n/a']
                   or any([e in s for e in ['tbc by', 'there is no']]))
        if illegal:
            printerr('{} is illegal'.format(i))
            # report['error'] += 1 # cound as an error?
        return illegal
    
    return (not (i is None
                 or i[CONCEPT] == "None"
                 or i[CODE] == "None"
                 or i[DESIGNATION] == "None"
                 or len(i[CODE]) == 0
                 or i[CONCEPT] is None
                 or i[CODE] is None
                 or any([is_illegal(e) for e in [i[CONCEPT], i[CODE], i[CODE_SYSTEM]]])))
    # print(i,"~~~~>",valid)
    # return valid


# if the key is already present, append the item to d[key],
# otherwise create d[key]
# this step does not group by site, so repetitions are ok
def insert_item(d, item):
    try:
        d[item[CONCEPT]].append(item)
    except KeyError:
        d[item[CONCEPT]] = [item]


def group_by(l, key):
    keys = set([x[key] for x in l])
    return({x: [y for y in l if y[key] == x and len(y[CODE])] for x in keys})


def importFile(f, config):
    printcolor("importing {}".format(f), color=bcolors.OKBLUE)
    d = {}
    wb = load_workbook(f)
    for sheetName in config['sheets']:
        sheet = wb[sheetName]
        for c in [a.upper() for a in config['columns']]:
            for r in range(3, sheet.max_row+1):
                # print(c, r)
                coord = Coord(c, r)
                i = build_item(wb, sheet, startcoord=coord)
                # print('item:', i)
                if item_is_valid(i):
                    assert i[CODE_SYSTEM] in [e[CODE_SYSTEM] for e in codesystems], 'error: <{}> not in {}'.format(i[CODE_SYSTEM], [e[CODE_SYSTEM] for e in codesystems])
                    i[CONCEPT] = sheetName + "|" + i[CONCEPT]
                    insert_item(d, i)
    # several codes for each site
    d = {k: group_by(v, SITE) for k, v in d.items()}
    process_items(d)


# ### terminologies uris ### #
def upload_terminologies(f):
    printok('updating terminologies from: {}'.format(f))
    terminos = pd.read_csv(f)
    print(terminos)
    def upload(x):
        printinfo('uploading terminology: {}'.format(x))
        resp = sendrequest('code-systems', method='post', data = x)
        print(resp)
    for x in terminos.iterrows():
        x = x[1]
        o = {CODE_SYSTEM: x[CODE_SYSTEM],
             CODE_SYSTEM_URI: x[CODE_SYSTEM_URI]}
        upload(o)

# ### Main ### #


if __name__ == "__main__":
    args = sys.argv
    global interactive
    global codesystems
    global dryrun
    global report
    global baseurl
    
    report = Counter(
        identical=0,
        different=0,
        new=0,
        error=0
    )

    interactive = '--interactive' in args
    dryrun = '--dry-run' in args
    FORCE = '--force' in args
    if '--url' in args:
        baseurl = args[args.index('--url')+1]
        args.remove('--url')
        args.remove(baseurl)
        printinfo('Using "{}" as base url'.format(baseurl))
    if FORCE:
        args.remove('--force')
    if dryrun:
        args.remove('--dry-run')
    if interactive:
        args.remove('--interactive')
    if '--NUKE' in args:
        args.remove('--NUKE')
        ans = sendrequest('all', method='delete')
        printwarn('deleting everything: {}'.format(ans))
    
    # print(mappings)

    if len(args) < 2:
        raise(Exception("please provide the path to a yaml configuration file"))
    else:
        configpath = args[1]
        dirname = os.path.dirname(configpath)
        def fullpath(e):
            return os.path.join(dirname, e)
        
        with open(configpath) as f:
            y = yaml.load(f)
            # print(y, 'terminologies' in y.keys(), y.keys())
            if 'terminologies' in y.keys():
                upload_terminologies(fullpath(y['terminologies']))

            jsondata = json.loads(sendrequest(url="all"))['data']
            mappings = jsondata['mappings']
            codesystems = jsondata['code_systems']

            [importFile(fullpath(k), e)
             for k, e in y['mappings'].items()]

    print('done.')
    print('──────────────────────────────')
    for category, count in report.items():
        print('{}: {}'.format(category, count))


# def tests_tmp():

# # [str(resolveCell(df, df['Observation & Units'], Coord(l,1))) for l in 'ABCDEFGHIJKLMNOPQRSTUV']

# f = '../data/data_2019_02_07/SemMapper-CDSMCodeMappings-with-cdsm-profile-v2-7Feb2019.xlsx'
# df = load_workbook(f)

# # for sh in config['sheets']:
# for sh in ["C3DP_CDSM_PROFILE", "Observations", "Conditions", "Medications", "Family History", "Procedures", "Observations & Units"]:
#     sheet = pd.DataFrame(df[sh].values)

    

#     def normalize_fused_cells(sheet, row):
#         col = ''
#         for ci in sheet.columns:
#             if sheet.iloc[row,ci] is None:
#                 sheet.iloc[row,ci] = col
#             else:
#                 col = sheet.iloc[row,ci]
    
